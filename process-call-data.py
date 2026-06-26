import sys
import os
import csv
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook


def should_ignore(sheet_name):
    ignored_prefixes = ("Sheet", "Documentation", "Report")
    return sheet_name.startswith(ignored_prefixes)


def find_target_sheet(workbook):
    all_sheets = workbook.sheetnames
    remaining_sheets = [s for s in all_sheets if not should_ignore(s)]

    data_sheets = [s for s in remaining_sheets if s.startswith("Data")]
    if data_sheets:
        return data_sheets[0]

    if len(remaining_sheets) == 1:
        return remaining_sheets[0]

    return None


def parse_datetime(value):
    if isinstance(value, datetime):
        return value
    return datetime.strptime(str(value).strip(), "%m.%d.%Y %H:%M:%S")


def parse_date(value):
    # Used for reading the date column from an existing CSV output.
    return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()


def read_existing_csv(output_csv):
    """Read a previously-generated output CSV if it exists.

    Returns a tuple (existing_result, existing_sites) where:
      - existing_result[day][site] is a list of int values previously recorded
        for that (day, site) pair. A list is used because past conflicts may
        have produced multiple rows for the same date.
      - existing_sites is the set of site column names from the header.
    """
    existing_result = defaultdict(lambda: defaultdict(list))
    existing_sites = set()

    if not os.path.exists(output_csv):
        return existing_result, existing_sites

    try:
        with open(output_csv, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None or len(header) < 2:
                return existing_result, existing_sites

            sites = [h.strip() for h in header[1:]]
            existing_sites.update(sites)

            for row_num, row in enumerate(reader, start=2):
                if not row or not row[0].strip():
                    # Skip rows without a date (defensive).
                    continue
                try:
                    day = parse_date(row[0])
                except Exception as e:
                    print(f"Warning: Skipping invalid existing CSV row {row_num}: {row} ({e})")
                    continue

                for i, site in enumerate(sites, start=1):
                    if i >= len(row):
                        break
                    cell = row[i].strip() if row[i] is not None else ""
                    if cell == "":
                        continue
                    try:
                        val = int(float(cell))
                    except Exception as e:
                        print(f"Warning: Skipping invalid existing value at row {row_num}, site '{site}': {cell} ({e})")
                        continue
                    existing_result[day][site].append(val)
    except Exception as e:
        print(f"Warning: Could not read existing output file '{output_csv}': {e}")
        return defaultdict(lambda: defaultdict(list)), set()

    return existing_result, existing_sites


def main():
    if len(sys.argv) != 2:
        print(f"Usage: python {sys.argv[0]} <xlsx_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    path = os.path.dirname(file_path)

    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error opening workbook: {e}")
        sys.exit(1)

    sheet_name = find_target_sheet(workbook)
    if not sheet_name:
        print("Error: Could not determine the correct sheet to process.")
        sys.exit(1)

    sheet = workbook[sheet_name]
    sheet.reset_dimensions()   # forces openpyxl to recalculate
    sheet.calculate_dimension(force=True)
    if sheet.max_column != 3:
        print(f"Warning: Sheet '{sheet_name}' does not have exactly 3 columns (found {sheet.max_column}).")
        #sys.exit(1)

    rows = sheet.iter_rows(values_only=True)

    header = next(rows, None)
    if header is None:
        print(f"Error: Sheet '{sheet_name}' is empty.")
        sys.exit(1)

    second_col_header = str(header[1]).strip()
    outputfile = second_col_header.split()[0].lower()
    output_csv = f"{path}/{outputfile}.csv"

    # --- Step 1: read previously-generated output (if any) ---
    existing_result, existing_sites = read_existing_csv(output_csv)

    # --- Step 2: process the xlsx into xlsx_result (max value per day+site) ---
    xlsx_result = defaultdict(dict)
    xlsx_sites = set()

    for row_num, row in enumerate(rows, start=2):
        if row is None or len(row) < 3:
            continue

        dt_val, site, data_val = row[0], row[1], row[2]

        if dt_val is None or site is None or data_val is None:
            continue

        try:
            dt = parse_datetime(dt_val)
            day = dt.date()
            site = str(site).strip()
            data_num = int(float(data_val))
        except Exception as e:
            print(f"Warning: Skipping invalid row {row_num}: {row} ({e})")
            continue

        xlsx_sites.add(site)

        current = xlsx_result[day].get(site)
        if current is None or data_num > current:
            xlsx_result[day][site] = data_num

    # --- Step 3: merge existing and xlsx data ---
    all_sites = sorted(existing_sites | xlsx_sites)
    all_days = sorted(set(existing_result.keys()) | set(xlsx_result.keys()))

    merged = defaultdict(lambda: defaultdict(list))
    conflict_count = 0

    for day in all_days:
        for site in all_sites:
            existing_vals = list(existing_result.get(day, {}).get(site, []))
            xlsx_val = xlsx_result.get(day, {}).get(site)

            if not existing_vals and xlsx_val is None:
                continue

            if xlsx_val is None:
                # Only existing has data - keep as-is.
                merged[day][site] = existing_vals
                continue

            if not existing_vals:
                # Only the xlsx has data.
                merged[day][site] = [xlsx_val]
                continue

            # Both sides have data.
            if xlsx_val in existing_vals:
                # Same data - keep only one copy (dedup against existing).
                # Preserve any existing duplicates already on disk untouched.
                merged[day][site] = existing_vals
            else:
                # Different - print conflict and keep both.
                print(
                    f"Conflict on {day.isoformat()} site={site}: "
                    f"existing={existing_vals}, new={xlsx_val}"
                )
                conflict_count += 1
                merged[day][site] = existing_vals + [xlsx_val]

    # --- Step 4: write merged output CSV ---
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["date"] + all_sites)

        for day in all_days:
            # Determine how many rows we need for this day (driven by the
            # site with the most stored values, e.g. multiple conflicting values).
            max_rows = 1
            for site in all_sites:
                vals = merged[day].get(site, [])
                if len(vals) > max_rows:
                    max_rows = len(vals)

            for i in range(max_rows):
                # Fill the date on every row so the file stays self-describing
                # and parseable on subsequent re-runs.
                out_row = [day.isoformat()]
                for site in all_sites:
                    vals = merged[day].get(site, [])
                    if i < len(vals):
                        out_row.append(vals[i])
                    else:
                        out_row.append("")
                writer.writerow(out_row)

    print(f"Processed sheet: {sheet_name}")
    print(f"Output written to: {output_csv}")
    print(f"Conflicts encountered: {conflict_count}")


if __name__ == "__main__":
    main()
